from io import BytesIO
from pathlib import Path

import fitz
import requests
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


OUTPUT_DIR = Path(r"C:\Users\15164\OneDrive\Desktop\Census Master Files\Florida Census")
OUTPUT_FILE = OUTPUT_DIR / "Alachua_County_Florida_1940_Census_Cities.xlsx"
PRECINCT_OUTPUT_FILE = OUTPUT_DIR / "Alachua_County_Florida_1940_Census_Precincts.xlsx"
IMAGE_DIR = OUTPUT_DIR / "Alachua_County_1940_Census_Images"
SOURCE_PDF_URL = "https://www2.census.gov/library/publications/decennial/1940/population-volume-1/33973538v1ch03.pdf"
SOURCE_DOCUMENT = "1940 Census of Population, Volume 1: Number of Inhabitants (Florida section)"
SOURCE_NOTE = (
    "Workbook generated from the official 1940 Census Bureau Florida population "
    "tables for incorporated places and county subdivisions."
)

# Values are taken from the 1940 Census of Population, Volume 1, Florida tables.
# 1940 values are the primary requested output. Earlier values are included where
# they were recoverable from the same report for convenience.
CITY_ROWS = [
    {
        "place": "Alachua",
        "county": "Alachua",
        "population_1940": 1081,
        "population_1930": 865,
        "population_1920": 778,
        "source_table": "Table 5 and Table 4",
        "source_page": "Florida p. 224-225",
        "source_spelling": "Alachua city",
    },
    {
        "place": "Archer",
        "county": "Alachua",
        "population_1940": 617,
        "population_1930": 576,
        "population_1920": 420,
        "source_table": "Table 4",
        "source_page": "Florida p. 224",
        "source_spelling": "Archer city",
    },
    {
        "place": "Gainesville",
        "county": "Alachua",
        "population_1940": 13757,
        "population_1930": 10465,
        "population_1920": 6860,
        "source_table": "Table 5 and Table 4",
        "source_page": "Florida p. 224-225",
        "source_spelling": "Gainesville city",
    },
    {
        "place": "Hawthorne",
        "county": "Alachua",
        "population_1940": 741,
        "population_1930": 600,
        "population_1920": 645,
        "source_table": "Table 5 and Table 4",
        "source_page": "Florida p. 224-225",
        "source_spelling": "Hawthorn town",
    },
    {
        "place": "High Springs",
        "county": "Alachua",
        "population_1940": 2010,
        "population_1930": 1864,
        "population_1920": 1719,
        "source_table": "Table 5 and Table 4",
        "source_page": "Florida p. 224-225",
        "source_spelling": "High Springs city",
    },
    {
        "place": "Micanopy",
        "county": "Alachua",
        "population_1940": 720,
        "population_1930": 725,
        "population_1920": 546,
        "source_table": "Table 5 and Table 4",
        "source_page": "Florida p. 224-225",
        "source_spelling": "Micanopy town",
    },
    {
        "place": "Newberry",
        "county": "Alachua",
        "population_1940": 785,
        "population_1930": 766,
        "population_1920": 917,
        "source_table": "Table 4",
        "source_page": "Florida p. 224",
        "source_spelling": "Newberry town",
    },
    {
        "place": "Waldo",
        "county": "Alachua",
        "population_1940": 667,
        "population_1930": 703,
        "population_1920": 571,
        "source_table": "Table 4",
        "source_page": "Florida p. 224",
        "source_spelling": "Waldo city",
    },
]

PRECINCT_ROWS = [
    {"precinct_number": 1, "precinct_name": "Waldo", "population_1940": 1100, "population_1930": 1488, "population_1920": 1181, "source_excerpt": "Pree. 1, Waldo / Waldo city"},
    {"precinct_number": 2, "precinct_name": "La Crosse", "population_1940": 639, "population_1930": 889, "population_1920": 1131, "source_excerpt": "Pree. 2, La Crosse / La Crosse town"},
    {"precinct_number": 3, "precinct_name": "Alachua", "population_1940": 2415, "population_1930": 2362, "population_1920": 2006, "source_excerpt": "Pree. 3, Alachua / Alachua city"},
    {"precinct_number": 4, "precinct_name": "Gainesville", "population_1940": 1947, "population_1930": None, "population_1920": None, "source_excerpt": "Pree. 4, Gainesville"},
    {"precinct_number": 5, "precinct_name": "Gainesville", "population_1940": 2591, "population_1930": None, "population_1920": None, "source_excerpt": "Pree. 5, Gainesville"},
    {"precinct_number": 6, "precinct_name": "Newberry", "population_1940": 2869, "population_1930": 2461, "population_1920": 3321, "source_excerpt": "Pree. 6, Newberry / Newberry town"},
    {"precinct_number": 7, "precinct_name": "Gainesville", "population_1940": 4991, "population_1930": 15388, "population_1920": 4277, "source_excerpt": "Pree. 7, Gainesville"},
    {"precinct_number": 8, "precinct_name": "Hawthorn", "population_1940": 1403, "population_1930": 1403, "population_1920": None, "source_excerpt": "Pree. 8, Hawthorn / Hawthorn town"},
    {"precinct_number": 9, "precinct_name": "Island Grove", "population_1940": 607, "population_1930": 582, "population_1920": 559, "source_excerpt": "Pree. 9, Island Grove"},
    {"precinct_number": 10, "precinct_name": "Micanopy", "population_1940": 1330, "population_1930": 1274, "population_1920": 1257, "source_excerpt": "Pree. 10, Micanopy / Micanopy town"},
    {"precinct_number": 11, "precinct_name": "Archer", "population_1940": 1805, "population_1930": 2015, "population_1920": 1866, "source_excerpt": "Pree. 11, Archer / Archer city"},
    {"precinct_number": 12, "precinct_name": "Arredondo", "population_1940": 611, "population_1930": 596, "population_1920": 551, "source_excerpt": "Pree. 12, Arredondo"},
    {"precinct_number": 13, "precinct_name": "Fairbanks", "population_1940": 352, "population_1930": 346, "population_1920": 380, "source_excerpt": "Pree. 13, Fairbanks"},
    {"precinct_number": 14, "precinct_name": "Melrose", "population_1940": 284, "population_1930": 354, "population_1920": 302, "source_excerpt": "Pree. 14, Melrose"},
    {"precinct_number": 15, "precinct_name": "Rochelle", "population_1940": 778, "population_1930": 755, "population_1920": 787, "source_excerpt": "Pree. 15, Rochelle"},
    {"precinct_number": 16, "precinct_name": "Evinston", "population_1940": 71, "population_1930": 52, "population_1920": 124, "source_excerpt": "Pree. 16, Evinston"},
    {"precinct_number": 17, "precinct_name": "Hague", "population_1940": 453, "population_1930": 526, "population_1920": 723, "source_excerpt": "Pree. 17, Hague"},
    {"precinct_number": 19, "precinct_name": "Orange Heights", "population_1940": 320, "population_1930": 311, "population_1920": 364, "source_excerpt": "Pree. 19, Orange Heights"},
    {"precinct_number": 20, "precinct_name": "High Springs", "population_1940": 1139, "population_1930": 2587, "population_1920": 2615, "source_excerpt": "Pree. 20, High Springs"},
    {"precinct_number": 21, "precinct_name": "High Springs", "population_1940": 1809, "population_1930": None, "population_1920": None, "source_excerpt": "Pree. 21, High Springs"},
    {"precinct_number": 22, "precinct_name": "Campville", "population_1940": 412, "population_1930": 427, "population_1920": 422, "source_excerpt": "Pree. 22, Campville"},
    {"precinct_number": 23, "precinct_name": "Gainesville", "population_1940": 2608, "population_1930": None, "population_1920": None, "source_excerpt": "Pree. 23, Gainesville"},
    {"precinct_number": 24, "precinct_name": "Bland", "population_1940": 1340, "population_1930": 763, "population_1920": None, "source_excerpt": "Pree. 24, Bland"},
    {"precinct_number": 25, "precinct_name": "Gainesville", "population_1940": 1161, "population_1930": 2358, "population_1920": 2997, "source_excerpt": "Pree. 25, Gainesville"},
    {"precinct_number": 26, "precinct_name": "Mounteocha", "population_1940": 637, "population_1930": 711, "population_1920": 930, "source_excerpt": "Pree. 26, Mounteocha"},
    {"precinct_number": 27, "precinct_name": "Gainesville", "population_1940": 1868, "population_1930": 13435, "population_1920": None, "source_excerpt": "Pree. 27, Gainesville"},
    {"precinct_number": 28, "precinct_name": "Gainesville", "population_1940": 3067, "population_1930": None, "population_1920": None, "source_excerpt": "Pree. 28, Gainesville"},
]

CITY_IMAGE_TARGETS = [
    {"place": "Alachua", "page_index": 150, "search_term": "Alachua"},
    {"place": "Archer", "page_index": 150, "search_term": "Archer"},
    {"place": "Gainesville", "page_index": 151, "search_term": "Gainesville"},
    {"place": "Hawthorne", "page_index": 151, "search_term": "Hawthorn"},
    {"place": "High Springs", "page_index": 151, "search_term": "High Springs"},
    {"place": "Micanopy", "page_index": 151, "search_term": "Micanopy"},
    {"place": "Newberry", "page_index": 151, "search_term": "Newberry"},
    {"place": "Waldo", "page_index": 151, "search_term": "Waldo"},
]


def autofit_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_index = column_cells[0].column
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 2, 40)


def add_header_row(worksheet, headers: list[str]) -> None:
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)


def create_city_detail_sheet(workbook: Workbook, row: dict) -> None:
    sheet = workbook.create_sheet(title=row["place"][:31])
    add_header_row(sheet, ["Field", "Value"])
    sheet.append(["Place", row["place"]])
    sheet.append(["County", row["county"]])
    sheet.append(["Population 1940", row["population_1940"]])
    sheet.append(["Population 1930", row["population_1930"]])
    sheet.append(["Population 1920", row["population_1920"]])
    sheet.append(["Source table", row["source_table"]])
    sheet.append(["Source page", row["source_page"]])
    sheet.append(["Source spelling in report", row["source_spelling"]])
    sheet.append(["Source document", SOURCE_DOCUMENT])
    sheet.append(["Note", SOURCE_NOTE])
    autofit_columns(sheet)


def create_single_city_workbook(row: dict) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = row["place"][:31]
    add_header_row(sheet, ["Field", "Value"])
    sheet.append(["Place", row["place"]])
    sheet.append(["County", row["county"]])
    sheet.append(["Population 1940", row["population_1940"]])
    sheet.append(["Population 1930", row["population_1930"]])
    sheet.append(["Population 1920", row["population_1920"]])
    sheet.append(["Source table", row["source_table"]])
    sheet.append(["Source page", row["source_page"]])
    sheet.append(["Source spelling in report", row["source_spelling"]])
    sheet.append(["Source document", SOURCE_DOCUMENT])
    sheet.append(["Note", SOURCE_NOTE])
    autofit_columns(sheet)
    return workbook


def build_precinct_workbook() -> Workbook:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Precincts"
    add_header_row(
        summary,
        [
            "Precinct Number",
            "Precinct Name",
            "Population 1940",
            "Population 1930",
            "Population 1920",
            "Source Excerpt",
            "Source Table",
            "Source Page",
            "Source Document",
        ],
    )

    for row in PRECINCT_ROWS:
        summary.append(
            [
                row["precinct_number"],
                row["precinct_name"],
                row["population_1940"],
                row["population_1930"],
                row["population_1920"],
                row["source_excerpt"],
                "Table 4",
                "Florida p. 224",
                SOURCE_DOCUMENT,
            ]
        )

    notes = workbook.create_sheet(title="Notes")
    add_header_row(notes, ["Field", "Value"])
    notes.append(["County", "Alachua County, Florida"])
    notes.append(["County total 1940", 38607])
    notes.append(["County total 1930", 34365])
    notes.append(["County total 1920", 31689])
    notes.append(["Source table", "Table 4"])
    notes.append(["Source page", "Florida p. 224"])
    notes.append(["Source document", SOURCE_DOCUMENT])
    notes.append([
        "Extraction note",
        "Some 1930 and 1920 values on the source page run together across newspaper-style columns in PDF text extraction. Rows are transcribed from the official table slice used to generate this workbook.",
    ])
    autofit_columns(summary)
    autofit_columns(notes)
    return workbook


def save_single_city_workbooks() -> list[Path]:
    output_paths = []
    for row in CITY_ROWS:
        workbook = create_single_city_workbook(row)
        output_path = OUTPUT_DIR / f"{row['place']}_Florida_1940_Census.xlsx"
        saved_path = save_workbook_with_fallback(workbook, output_path)
        output_paths.append(saved_path)
    return output_paths


def fetch_source_pdf() -> fitz.Document:
    response = requests.get(SOURCE_PDF_URL, timeout=60)
    response.raise_for_status()
    return fitz.open(stream=BytesIO(response.content), filetype="pdf")


def save_workbook_with_fallback(workbook: Workbook, output_path: Path) -> Path:
    try:
        workbook.save(output_path)
        return output_path
    except PermissionError:
        fallback_path = output_path.with_stem(f"{output_path.stem}_updated")
        workbook.save(fallback_path)
        return fallback_path


def save_page_image(page: fitz.Page, output_path: Path, clip: fitz.Rect | None = None) -> None:
    matrix = fitz.Matrix(2, 2)
    pixmap = page.get_pixmap(matrix=matrix, clip=clip, alpha=False)
    pixmap.save(output_path)


def export_source_images() -> list[Path]:
    IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    document = fetch_source_pdf()
    saved_paths: list[Path] = []

    page_224 = document[150]
    page_225 = document[151]
    page_224_path = IMAGE_DIR / "Florida_1940_Census_p224_Alachua_Precincts.png"
    page_225_path = IMAGE_DIR / "Florida_1940_Census_p225_Alachua_Cities.png"
    save_page_image(page_224, page_224_path)
    save_page_image(page_225, page_225_path)
    saved_paths.extend([page_224_path, page_225_path])

    for target in CITY_IMAGE_TARGETS:
        page = document[target["page_index"]]
        matches = page.search_for(target["search_term"])
        if not matches:
            continue
        match = matches[0]
        clip = fitz.Rect(0, max(match.y0 - 16, 0), page.rect.width, min(match.y1 + 18, page.rect.height))
        image_path = IMAGE_DIR / f"{target['place'].replace(' ', '_')}_row.png"
        save_page_image(page, image_path, clip=clip)
        saved_paths.append(image_path)

    return saved_paths


def build_workbook() -> Workbook:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"

    add_header_row(
        summary,
        [
            "Place",
            "County",
            "Population 1940",
            "Population 1930",
            "Population 1920",
            "Source Table",
            "Source Page",
            "Source Spelling",
            "Source Document",
        ],
    )

    for row in CITY_ROWS:
        summary.append(
            [
                row["place"],
                row["county"],
                row["population_1940"],
                row["population_1930"],
                row["population_1920"],
                row["source_table"],
                row["source_page"],
                row["source_spelling"],
                SOURCE_DOCUMENT,
            ]
        )
        create_city_detail_sheet(workbook, row)

    autofit_columns(summary)
    return workbook


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook()
    combined_path = save_workbook_with_fallback(workbook, OUTPUT_FILE)
    precinct_workbook = build_precinct_workbook()
    precinct_path = save_workbook_with_fallback(precinct_workbook, PRECINCT_OUTPUT_FILE)
    city_files = save_single_city_workbooks()
    image_files = export_source_images()
    print(f"Created {combined_path}")
    print(f"Created {precinct_path}")
    print("Places:")
    for row in CITY_ROWS:
        print(f"- {row['place']}: {row['population_1940']}")
    print("Single-city workbooks:")
    for path in city_files:
        print(f"- {path.name}")
    print("Image exports:")
    for path in image_files:
        print(f"- {path.name}")


if __name__ == "__main__":
    main()